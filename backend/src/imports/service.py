"""
Import Excel back-office (CDC §3.6).

Sécurité :
- Extension ET signature de fichier vérifiées (pas de confiance sur le
  Content-Type déclaré par le client).
- Taille bornée avant tout parsing (évite un DoS mémoire via fichier énorme).
- Nombre de lignes plafonné (`MAX_IMPORT_ROWS`).
- Lecture en mode `data_only=True` : on récupère les valeurs mises en cache
  par Excel, jamais les formules elles-mêmes — aucune formule n'est
  évaluée côté serveur (openpyxl n'exécute jamais de formule de toute façon,
  mais `data_only=True` évite même de manipuler la chaîne de formule brute).
- Chaque ligne est validée indépendamment via les mêmes contraintes
  Pydantic que la création manuelle (`RobotCreate`) : aucun contournement
  de la validation métier via l'import de masse.
"""
from __future__ import annotations

import uuid
from datetime import datetime, timezone
from io import BytesIO

from fastapi import UploadFile
from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy.ext.asyncio import AsyncSession

from src.config import settings
from src.imports import models, schemas
from src.robots import models as robot_models, schemas as robot_schemas
from src.robots.service import create_robot
from src.shared.enums import ImportRowStatus
from src.shared.exceptions import ConflictError, InvalidFileError

EXPECTED_COLUMNS = [
    "ID_Robot", "Modèle", "Type", "Payload_kg", "Rayon_mm", "Axes", "Type_baie",
    "Annee_service", "Heures_utilisation", "Quantite_stock", "Prix_catalogue_EUR",
    "Remise_pct", "Prestations_dispo", "Protection_IP", "Description_courte",
]

_MONTAGE_PAR_DEFAUT = "À définir (compléter manuellement)"


def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


async def _read_workbook(file: UploadFile) -> BytesIO:
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise InvalidFileError("Seuls les fichiers Excel (.xlsx) sont acceptés.")

    raw = await file.read()
    max_bytes = settings.MAX_UPLOAD_SIZE_MB * 1024 * 1024
    if len(raw) > max_bytes:
        raise InvalidFileError(f"Fichier trop volumineux (max {settings.MAX_UPLOAD_SIZE_MB} Mo).")
    if len(raw) == 0:
        raise InvalidFileError("Fichier vide.")

    # Signature ZIP (xlsx = zip) — rejette un fichier renommé qui n'est pas un vrai xlsx.
    if raw[:2] != b"PK":
        raise InvalidFileError("Le fichier n'est pas un classeur Excel (.xlsx) valide.")

    return BytesIO(raw)


def _row_to_robot_create(row_values: dict[str, object], row_number: int) -> robot_schemas.RobotCreate:
    prestations_raw = str(row_values.get("Prestations_dispo") or "").strip()
    prestations = [
        robot_schemas.PrestationSchema(niveau=n.strip(), garantie_mois=None)
        for n in prestations_raw.split(";")
        if n.strip()
    ]

    return robot_schemas.RobotCreate(
        id=str(row_values.get("ID_Robot") or "").strip(),
        modele=str(row_values.get("Modèle") or "").strip(),
        type=str(row_values.get("Type") or "").strip(),
        categorie=str(row_values.get("Type") or "").strip(),  # pas de colonne dédiée : reprend le Type
        anneeMiseEnService=int(row_values.get("Annee_service") or 0),
        heuresUtilisation=int(row_values.get("Heures_utilisation") or 0),
        descriptionCourte=str(row_values.get("Description_courte") or "").strip(),
        caracteristiques=robot_schemas.CaracteristiquesSchema(
            payloadKg=float(row_values.get("Payload_kg") or 0),
            rayonActionMm=int(row_values.get("Rayon_mm") or 0),
            axes=int(row_values.get("Axes") or 0),
            typeBaie=str(row_values.get("Type_baie") or "").strip(),
            protectionIp=str(row_values.get("Protection_IP") or "").strip(),
            montage=_MONTAGE_PAR_DEFAUT,
        ),
        prestationsDisponibles=prestations,
        prixCatalogueEUR=float(row_values.get("Prix_catalogue_EUR") or 0),
        remisePct=float(row_values.get("Remise_pct") or 0),
        quantiteStock=int(row_values.get("Quantite_stock") or 0),
        nombreOffresEnCours=0,
    )


async def process_import(
    db: AsyncSession, file: UploadFile, operator_id: str
) -> schemas.ImportReport:
    buffer = await _read_workbook(file)

    try:
        workbook = load_workbook(buffer, read_only=True, data_only=True)
    except Exception as exc:  # fichier corrompu, mot de passe, etc.
        raise InvalidFileError("Impossible de lire le fichier Excel (corrompu ou protégé).") from exc

    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)

    try:
        header = next(rows_iter)
    except StopIteration:
        raise InvalidFileError("Le fichier est vide.")

    header_map = {str(h).strip(): idx for idx, h in enumerate(header) if h is not None}
    missing = [c for c in EXPECTED_COLUMNS if c not in header_map]
    if missing:
        raise InvalidFileError(
            "Colonnes manquantes dans le fichier : " + ", ".join(missing)
        )

    details: list[models.ImportRowResult] = []
    accepte = erreurs = doublons = 0
    seen_ids: set[str] = set()

    for row_number, raw_row in enumerate(rows_iter, start=2):
        if row_number - 1 > settings.MAX_IMPORT_ROWS:
            details.append(
                models.ImportRowResult(
                    ligne=row_number, id_robot="—", statut=ImportRowStatus.ERREUR,
                    message=f"Import limité à {settings.MAX_IMPORT_ROWS} lignes.",
                )
            )
            erreurs += 1
            break

        if raw_row is None or all(v is None for v in raw_row):
            continue  # ligne vide ignorée silencieusement

        row_values = {col: raw_row[idx] if idx < len(raw_row) else None for col, idx in header_map.items()}
        robot_id = str(row_values.get("ID_Robot") or "").strip()

        try:
            payload = _row_to_robot_create(row_values, row_number)
        except (ValidationError, ValueError, TypeError) as exc:
            details.append(
                models.ImportRowResult(
                    ligne=row_number, id_robot=robot_id or "—",
                    statut=ImportRowStatus.ERREUR, message=str(exc)[:500],
                )
            )
            erreurs += 1
            continue

        if payload.id in seen_ids:
            details.append(
                models.ImportRowResult(
                    ligne=row_number, id_robot=payload.id,
                    statut=ImportRowStatus.DOUBLON, message="ID en doublon dans le fichier importé.",
                )
            )
            doublons += 1
            continue
        seen_ids.add(payload.id)

        existing = await db.get(robot_models.Robot, payload.id)
        if existing is not None:
            details.append(
                models.ImportRowResult(
                    ligne=row_number, id_robot=payload.id,
                    statut=ImportRowStatus.DOUBLON, message="Une fiche avec cet ID existe déjà en base.",
                )
            )
            doublons += 1
            continue

        try:
            await create_robot(db, payload)
        except ConflictError as exc:
            details.append(
                models.ImportRowResult(
                    ligne=row_number, id_robot=payload.id, statut=ImportRowStatus.DOUBLON, message=str(exc),
                )
            )
            doublons += 1
            continue

        details.append(
            models.ImportRowResult(ligne=row_number, id_robot=payload.id, statut=ImportRowStatus.ACCEPTE)
        )
        accepte += 1

    batch = models.ImportBatch(
        id=str(uuid.uuid4()),
        operator_id=operator_id,
        filename=file.filename or "import.xlsx",
        total_lignes=accepte + erreurs + doublons,
        accepte=accepte,
        erreurs=erreurs,
        doublons=doublons,
        details=details,
    )
    db.add(batch)
    await db.commit()

    return schemas.ImportReport(
        total_lignes=batch.total_lignes,
        accepte=accepte,
        erreurs=erreurs,
        doublons=doublons,
        details=[
            schemas.ImportRowResultSchema(
                ligne=d.ligne, id_robot=d.id_robot, statut=d.statut, message=d.message
            )
            for d in details
        ],
    )
